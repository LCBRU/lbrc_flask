import xlwt
from openpyxl import Workbook
from lbrc_flask.column_data import CsvData, Excel97Data, ExcelData


def test__column_data__csv(tmp_path):
    csv_file = tmp_path / "test.csv"
    csv_file.write_text('fred,mary,ed\nsoup,fish,2000-01-02,pie')

    out = CsvData(csv_file)

    assert out.get_column_names() == ['fred', 'mary', 'ed', 'column 3']

    d = list(out.iter_rows())
    assert len(d) == 1

    for x in out.iter_rows():
        assert x['fred'] == 'soup'
        assert x['mary'] == 'fish'
        assert x['ed'] == '2000-01-02'
        assert x['column 3'] == 'pie'


def test__column_data__xls(tmp_path):
    xls_file = tmp_path / "test.xls"

    wb = xlwt.Workbook()
    ws = wb.add_sheet('A Test Sheet')

    ws.write(0, 0, 'fred')
    ws.write(0, 1, 'mary')
    ws.write(0, 2, 'ed')
    ws.write(1, 0, 'soup')
    ws.write(1, 1, 'fish')
    ws.write(1, 2, '2000-01-02')
    ws.write(1, 3, 'pie')

    wb.save(xls_file)

    out = Excel97Data(xls_file)

    assert out.get_column_names() == ['fred', 'mary', 'ed']

    d = list(out.iter_rows())
    assert len(d) == 1

    for x in out.iter_rows():
        assert x['fred'] == 'soup'
        assert x['mary'] == 'fish'
        assert x['ed'] == '2000-01-02'


def test__column_data__xlsx(tmp_path):
    xlsx_file = tmp_path / "test.xlsx"

    wb = Workbook()
    ws = wb.active

    ws.cell(column=1, row=1, value='fred')
    ws.cell(column=2, row=1, value='mary')
    ws.cell(column=3, row=1, value='ed')
    ws.cell(column=1, row=2, value='soup')
    ws.cell(column=2, row=2, value='fish')
    ws.cell(column=3, row=2, value='2000-01-02')
    wb.save(xlsx_file)

    out = ExcelData(xlsx_file)

    assert out.get_column_names() == ['fred', 'mary', 'ed']

    d = list(out.iter_rows())
    assert len(d) == 1

    for x in out.iter_rows():
        assert x['fred'] == 'soup'
        assert x['mary'] == 'fish'
        assert x['ed'] == '2000-01-02'
