import chardet
import csv
import xlrd
from openpyxl import load_workbook
from itertools import takewhile, zip_longest
from itertools import islice


class ColumnData():
    def __init__(self, filepath):
        self.filepath = filepath


class ExcelData(ColumnData):

    def get_column_names(self):
        wb = load_workbook(filename=self.filepath, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(min_row=1, max_row=1)
        first_row = next(rows)

        return [c.value for c in takewhile(lambda x: x.value, first_row)]

    def iter_rows(self):
        wb = load_workbook(filename=self.filepath, read_only=True)
        ws = wb.active

        column_names = self.get_column_names()
        for r in ws.iter_rows(min_row=2, values_only=True):
            yield dict(zip(column_names, r))


class Excel97Data(ColumnData):

    def get_column_names(self):
        wb = xlrd.open_workbook(filename=self.filepath)
        ws = wb.sheet_by_index(0)
        first_row = ws.row(0)

        return [c.value for c in takewhile(lambda x: x.value, first_row)]

    def iter_rows(self):
        wb = xlrd.open_workbook(filename=self.filepath, formatting_info=True)
        ws = wb.sheet_by_index(0)

        column_names = self.get_column_names()
        rows = ws.get_rows()
        next(rows)
        for r in rows:
            yield dict(zip(column_names, [self._value_from_cell(c, wb.datemode) for c in r]))

    def _value_from_cell(self, cell, datemode):
        if cell.ctype == xlrd.book.XL_CELL_DATE:
            return xlrd.xldate.xldate_as_datetime(cell.value, datemode)
        else:
            return cell.value


class CsvData(ColumnData):

    def get_column_names(self):
        result = []

        with open(self.filepath, 'r', encoding=self._get_encoding()) as f:
            d_reader = csv.DictReader(f)

            result = d_reader.fieldnames

            for row in d_reader:
                for i in range(len(result), len(row)):
                    result.append(f'Column {i}')

        return result 

    def _get_encoding(self):
        rawdata = open(self.filepath, 'rb').read()
        result = chardet.detect(rawdata)
        return result['encoding']

    def iter_rows(self):
        column_names = self.get_column_names()

        with open(self.filepath, 'r', encoding=self._get_encoding()) as f:
            reader = csv.reader(f)

            for row in islice(reader, 1, None):
                yield dict(zip_longest(column_names, row, fillvalue=''))
