from flask import send_file, render_template
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from tempfile import NamedTemporaryFile
from datetime import datetime
from flask_weasyprint import HTML, render_pdf


def excel_download(title, headers, details):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    bold = Font(bold=True)

    for col, h in enumerate(headers, start=1):
        ws['{}1'.format(get_column_letter(col))].style = 'Headline 1'
        ws.cell(column=col, row=1, value=h.title())

    for row, d in enumerate(details, start=2):
        for col, h in enumerate(headers, start=1):
            ws.cell(column=col, row=row, value=d.get(h, ''))

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.flush()
        return send_file(
            tmp.name,
            as_attachment=True,
            attachment_filename='{}_{}.xlsx'.format(title, datetime.utcnow().strftime("%Y%m%d_%H%M%S")),
            cache_timeout=0,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )


def pdf_download(template, title="report", **kwargs):
    resp = render_pdf(HTML(string=render_template(template, **kwargs)))
    resp.headers['Content-Disposition'] = f'attachment;filename={title}_{datetime.utcnow():%Y%m%d_%H%M%S}.pdf'
    return resp