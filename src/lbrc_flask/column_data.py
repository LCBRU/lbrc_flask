from dataclasses import dataclass
import string
from typing import Optional
import chardet
import csv
import xlrd
from openpyxl import load_workbook
from itertools import takewhile, zip_longest, compress, islice
from lbrc_flask.validators import is_integer, is_number, parse_date_or_none
from lbrc_flask.lookups import Lookup, LookupRepository


class ColumnData():
    def __init__(self, filepath, header_rows: int = 1, column_header_row: int = 1):
        self.filepath = filepath
        self.header_rows = header_rows
        self.column_header_row = column_header_row


class ExcelData(ColumnData):
    def __init__(self, filepath, header_rows: int = 1, worksheet=None, column_header_row: int = 1):
        super().__init__(filepath, header_rows, column_header_row=column_header_row)
        self.worksheet = worksheet

    def get_worksheet(self):
        wb = self.get_workbook()

        if self.worksheet is None:
            return wb.active
        else:
            return wb[self.worksheet]

    def get_workbook(self):
        return load_workbook(filename=self.filepath, read_only=True)
    
    def has_worksheet(self):
        return self.worksheet in self.get_workbook().sheetnames

    def get_column_names(self):
        rows = self.get_worksheet().iter_rows(min_row=self.column_header_row, max_row=self.column_header_row)
        header_row = next(rows)

        return [str(c.value).lower() for c in takewhile(lambda x: x.value, header_row)]

    def iter_rows(self):
        column_names = self.get_column_names()
        for r in self.get_worksheet().iter_rows(min_row=self.header_rows + 1, values_only=True):
            yield dict(zip(column_names, r))


class Excel97Data(ColumnData):

    def get_column_names(self):
        wb = xlrd.open_workbook(filename=self.filepath)
        ws = wb.sheet_by_index(0)
        header_row = ws.row(self.column_header_row)

        return [c.value.lower() for c in takewhile(lambda x: x.value, header_row)]

    def iter_rows(self):
        wb = xlrd.open_workbook(filename=self.filepath, formatting_info=True)
        ws = wb.sheet_by_index(0)

        column_names = self.get_column_names()
        rows = ws.get_rows()
        for _ in range(self.header_rows):
            next(rows)
        for r in rows:
            yield dict(zip(column_names, [self._value_from_cell(c, wb.datemode) for c in r]))

    def _value_from_cell(self, cell, datemode):
        if cell.ctype == xlrd.book.XL_CELL_DATE:
            return xlrd.xldate.xldate_as_datetime(cell.value, datemode)
        else:
            return cell.value


class CsvData(ColumnData):

    def get_column_names(self):
        result = []

        with open(self.filepath, 'r', encoding=self._get_encoding()) as f:
            d_reader = csv.DictReader(f)

            result = [cn.lower() for cn in d_reader.fieldnames]

            for row in d_reader:
                for i in range(len(result), len(row)):
                    result.append(f'column {i}')

        return result

    def _get_encoding(self):
        rawdata = open(self.filepath, 'rb').read()
        result = chardet.detect(rawdata)
        return result['encoding']

    def iter_rows(self):
        column_names = self.get_column_names()

        with open(self.filepath, 'r', encoding=self._get_encoding()) as f:
            reader = csv.reader(f)

            for row in islice(reader, self.header_rows, None):
                yield dict(zip_longest(column_names, row, fillvalue=''))


class ColumnsDefinition():
    @staticmethod
    def field_name_from_column_name(column_name: str):
        return column_name.lower().replace(' ', '_')

    @property
    def minimum_row_count(self):
        return None
    
    @property
    def maximum_row_count(self):
        return None
    
    @property
    def column_definition(self):
        return None
    
    @property
    def column_names(self):
        return [d.name for d in self.column_definition]

    def row_filter(self, spreadsheet):
        return self.rows_with_any_fields(spreadsheet)

    def definition_for_column_name(self, name):
        for d in self.column_definition:
            if d.name == name:
                return d

    def validation_errors(self, spreadsheet):
        errors = []

        errors.extend(self.column_validation_errors(spreadsheet))
        errors.extend(self.data_validation_errors(spreadsheet))
        errors.extend(self.row_validation_errors(spreadsheet))
        errors.extend(self.custom_validation_errors(spreadsheet))

        return errors
    
    def custom_validation_errors(self, spreadsheet):
        return []

    def column_validation_errors(self, spreadsheet):

        def column_is_missing(column_name):
            return not any(c for c in spreadsheet.get_column_names() if c.startswith(column_name.lower()))

        missing_columns = [c for c in self.column_names if column_is_missing(c)]

        return [ColumnsDefinitionValidationMessage(
            type=ColumnsDefinitionValidationMessage.TYPE__ERROR,
            message=f"Missing column '{m}'"
        ) for m in missing_columns]

    def row_validation_errors(self, spreadsheet):
        rows = len(list(self.iter_filtered_data(spreadsheet)))

        messages = []

        if (min := self.minimum_row_count) is not None:
            if min > rows:
                messages.append(f"Expected a minimum of {min} rows, but {rows} were found")

        if (max := self.maximum_row_count) is not None:
            if max < rows:
                messages.append(f"Expected a maximum of {max} rows, but {rows} were found")

        return [
            ColumnsDefinitionValidationMessage(
                type=ColumnsDefinitionValidationMessage.TYPE__ERROR,
                message=m
            )
        for m in messages]

    def data_validation_errors(self, spreadsheet):
        result = []

        for i, row in enumerate(self.iter_filtered_data(spreadsheet), 1):
            for e in self._field_errors_for_def(row):
                e.row = i
                result.append(e)

        return result
    
    def iter_filtered_data(self, spreadsheet):
        row_filter = self.row_filter(spreadsheet)

        if row_filter is None:
            return spreadsheet.iter_rows()
        else:
            return compress(spreadsheet.iter_rows(), row_filter)
    
    def translated_data(self, spreadsheet):
        for row in self.iter_filtered_data(spreadsheet):
            result = {}

            for cd in self.column_definition:
                result.update(cd.get_translated_data(row))
            
            yield result

    def rows_with_all_fields(self, spreadsheet):
        result = []

        for row in spreadsheet.iter_rows():
            result.append(all([d.has_value(row) or d.allow_null for d in self.column_definition]))

        return result

    def rows_with_any_fields(self, spreadsheet):
        result = []

        for row in spreadsheet.iter_rows():
            result.append(any([d.has_value(row) for d in self.column_definition]))

        return result

    def _field_errors_for_def(self, row: dict):
        result = []
        for col_def in self.column_definition:
            messages = col_def.validation_errors(row)
            if (messages):
                print(row)
            result.extend(messages)
        
        return result


@dataclass
class ColumnsDefinitionValidationMessage:
    TYPE__ERROR = 'Error'
    TYPE__WARNING = 'Warning'

    TYPE_NAMES = [
        TYPE__ERROR,
        TYPE__WARNING,
    ]

    type: str
    message: str
    row: Optional[int] = None

    @property
    def is_error(self):
        return self.type == ColumnsDefinitionValidationMessage.TYPE__ERROR

    def __post_init__(self):
        if self.type not in ColumnsDefinitionValidationMessage.TYPE_NAMES:
            raise ValueError(f"Type is '{self.type}' must be one of {', '.join(ColumnsDefinitionValidationMessage.TYPE_NAMES)}")


@dataclass
class ColumnDefinition:
    name: str
    allow_null: bool = False
    skip_errors: bool = False
    max_length: Optional[int] = None
    translated_name: Optional[str] = None

    def _strip(self, value):
        return value.strip()
    
    def is_defined_in_row(self, row: dict):
        for k,v in row.items():
            if k.startswith(self.name.lower()):
                return True
        
        return False

    def value(self, row: dict):
        result = None

        for k,v in row.items():
            if k.startswith(self.name.lower()):
                result = v
                break
        
        if str(result).startswith('='):
            result = None

        if isinstance(result, str):
            result = self._strip(result)
        
        if isinstance(result, str) and len(result) == 0 and self.allow_null:
            result = None

        return result

    def stringed_value(self, row: dict):
        val = self.value(row)

        if val is None:
            val = ''
        
        val = str(val)
        val = self._strip(val)

        return val

    def has_value(self, row: dict):
        return len(self.stringed_value(row)) > 0

    def get_translated_name(self):
        return self.translated_name or self.name

    def _type_validation_errors(self, row: dict):
        return []

    def validation_errors(self, row: dict):
        result = []

        if self.is_defined_in_row(row):
            if self.has_value(row):
                result.extend(self._type_validation_errors(row))
            elif not self.allow_null:
                result.append(self._create_message("Data is missing"))

        return result
    
    def _create_message(self, error):
        if self.skip_errors:
            message_type = ColumnsDefinitionValidationMessage.TYPE__WARNING
        else:
            message_type = ColumnsDefinitionValidationMessage.TYPE__ERROR

        return ColumnsDefinitionValidationMessage(
            type=message_type,
            message=f"{self.name}: {error}",
        )

    def get_object_value(self, obj):
        return getattr(obj, self.translated_name)
    
    def get_translated_data(self, row):
        return {self.translated_name: self.get_translated_value(row)}
    
    def get_translated_value(self, row):
        return self.value(row)


@dataclass
class StringColumnDefinition(ColumnDefinition):
    max_length: Optional[int] = None

    def _type_validation_errors(self, row: dict):
        result = []

        if max_length := self.max_length:
            actual_length = len(str(self.value(row)))
            if actual_length > max_length:
               result.append(self._create_message(f"Text is longer than {max_length} characters ({actual_length})"))
        
        return result
    

@dataclass
class NumericColumnDefinition(ColumnDefinition):
    def _strip(self, value):
        last_length = len(value)
        current_length = 0

        while last_length > current_length:
            last_length = current_length
            value = value.strip().strip('£$€¥').replace(',', '')
            current_length = len(value)

        return value
    
    def _type_validation_errors(self, row: dict):
        result = []

        if not is_number(self.stringed_value(row)):
            result.append(self._create_message(f"Invalid value of '{self.stringed_value(row)}'"))
        
        return result

    def get_translated_value(self, row):
        value = self.value(row)

        if value is None:
            return None

        if not is_number(self.stringed_value(row)):
            return None

        return float(value)

@dataclass
class IntegerColumnDefinition(NumericColumnDefinition):
    def _type_validation_errors(self, row: dict):
        result = []

        if not is_integer(self.stringed_value(row)):
            result.append(self._create_message(f"Invalid value of '{self.value(row)}'"))
        
        return result

    def get_translated_value(self, row):
        value = self.value(row)

        if value is None:
            return None

        if not is_integer(value):
            return None

        return int(value)


@dataclass
class DateColumnDefinition(ColumnDefinition):
    def _type_validation_errors(self, row: dict):
        result = []

        if parse_date_or_none(self.value(row)) is None:
            result.append(self._create_message(f"Invalid value of '{self.value(row)}'"))
        
        return result

    def get_translated_value(self, row):
        return parse_date_or_none(self.value(row))


@dataclass
class BooleanColumnDefinition(ColumnDefinition):
    TRUE_VALUES = ['y', 'yes', 'true']
    FALSE_VALUES = ['n', 'no', 'false']

    def _type_validation_errors(self, row: dict):
        result = []

        if not self.stringed_value(row).lower() in BooleanColumnDefinition.TRUE_VALUES + BooleanColumnDefinition.FALSE_VALUES:
            result.append(self._create_message(f"Invalid value of '{self.value(row)}'"))
        
        return result

    def get_translated_value(self, row):
        value = self.value(row)
        result = None

        if value is not None:
            if isinstance(value, bool):
                result = value
            elif value.lower() in BooleanColumnDefinition.TRUE_VALUES:
                result = True
            elif value.lower() in BooleanColumnDefinition.FALSE_VALUES:
                result = False
        
        return result

@dataclass
class LookupColumnDefinition(ColumnDefinition):
    def _strip(self, value):
        last_length = len(value)
        current_length = 0

        while last_length > current_length:
            last_length = current_length
            value = value.strip().strip('.,;')
            current_length = len(value)

        return value
    
    lookup_class: Optional[Lookup] = None

    def _get_lookup(self, row: dict):
        return LookupRepository(self.lookup_class).get(self.stringed_value(row))

    def _type_validation_errors(self, row: dict):
        result = []

        if self._get_lookup(row) is None:
            result.append(self._create_message(f"Does not exist ('{self.stringed_value(row)}')"))

        return result

    def get_object_value(self, obj):
        value = super().get_object_value(obj)

        if value:
            return value.name

        return None

    def get_translated_data(self, row: dict):
        result = {self.translated_name: self.value(row)}

        lookup = self._get_lookup(row)

        if lookup is None:
            result[f'{self.translated_name}_id'] = None
        else:
            result[f'{self.translated_name}_id'] = lookup.id
        
        return result
