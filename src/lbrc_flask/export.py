import csv
from fileinput import filename
from flask import send_file, render_template
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from tempfile import NamedTemporaryFile
import datetime
from flask_weasyprint import HTML, render_pdf


def excel_download(title, headers, details):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    bold = Font(bold=True)

    for col, h in enumerate(headers, start=1):
        ws['{}1'.format(get_column_letter(col))].style = 'Headline 1'
        ws.cell(column=col, row=1, value=h.title())

    for row, d in enumerate(details, start=2):
        for col, h in enumerate(headers, start=1):
            ws.cell(column=col, row=row, value=d.get(h, ''))

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.flush()
        return send_file(
            tmp.name,
            as_attachment=True,
            download_name='{}_{}.xlsx'.format(title, datetime.datetime.now(datetime.UTC).strftime("%Y%m%d_%H%M%S")),
            max_age=0,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )


def csv_download(title, headers, details):
    with NamedTemporaryFile(mode='w+', delete=True, encoding='utf-8') as tmp:
        writer = csv.DictWriter(tmp, fieldnames=list(headers))

        writer.writeheader()

        for d in details:
            writer.writerow(d)

        tmp.flush()
        tmp.seek(0)

        return send_file(
            tmp.name,
            as_attachment=True,
            download_name=f'{title}_{datetime.datetime.now(datetime.UTC):%Y%m%d_%H%M%S}.csv',
        )


def pdf_download(template, title="report", path=None, **kwargs):
    with NamedTemporaryFile(mode='w+', delete=True, encoding='utf-8') as tmp_html, NamedTemporaryFile(mode='w+', delete=True) as tmp_pdf:
        tmp_html.write(render_template(template, **kwargs))
        tmp_html.flush()
        tmp_html.seek(0)

        weasy_html = HTML(filename=tmp_html.name, base_url=path)
        weasy_html.write_pdf(tmp_pdf.name)

        tmp_pdf.flush()
        tmp_pdf.seek(0)

        return send_file(
            tmp_pdf.name,
            as_attachment=True,
            download_name=f'{title}_{datetime.datetime.now(datetime.UTC):%Y%m%d_%H%M%S}.pdf',
        )
