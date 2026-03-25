import csv
import io
import xlwt
from openpyxl import Workbook
from faker.providers import BaseProvider
from faker import Faker
from typing import Optional
from functools import cache


class FakeXlsxWorksheet:
    def __init__(self, name: str, headers, data, headers_on_row=1):
        self.name = name
        self.headers = headers
        self.data = data
        self.headers_on_row = headers_on_row

    def create_worksheet(self, workbook: Workbook):
        if self.name is None:
            ws1 = workbook.active
        else:
            ws1 = workbook.create_sheet(self.name)

        for _ in range(1, self.headers_on_row):
            ws1.append([])

        ws1.append(list(self.headers))

        for d in self.data:
            row = []
            for h in self.headers:
                row.append(d.get(h.lower(), ''))
            ws1.append(row)


class FakeXlsxFile():
    def __init__(self, filename: str):
        self.filename = filename
        self.worksheets: list[FakeXlsxWorksheet] = []

    def get_iostream(self):
        result = io.BytesIO()
        self.save(result)

        return result.getvalue()

    def save(self, filepath):
        workbook= Workbook()

        for ws in self.worksheets:
            ws.create_worksheet(workbook)

        workbook.save(filepath)

    def add_worksheet(self, name: str, headers: Optional[list[str]]=None, data: Optional[dict]=None, headers_on_row: int=1):
        self.worksheets.append(FakeXlsxWorksheet(
            name=name,
            headers=headers,
            data=data,
            headers_on_row=headers_on_row,
        ))


class XlsxCreator:
    def __init__(self, provider: BaseProvider):
        self.provider: BaseProvider = provider
        self.populated_with_defaults: bool = False
        # The line below causes unique to fail because a
        # different faker is created each time, despite being for the
        # same generator.  Use singleton instead?
        self.faker: Faker = Faker("en_GB", generator=provider.generator)

    def get(self, headers: Optional[list[str]]=None, data: Optional[dict]=None, filename: Optional[str]=None, headers_on_row: int=1):
        headers = list(headers)
        filename = filename or self.faker.file_name(extension='xlsx')

        result = FakeXlsxFile(filename=filename)
        result.add_worksheet(
            name=None,
            headers=headers,
            data=data,
            headers_on_row=headers_on_row,
        )

        return result


class FakeCsvFile():
    def __init__(self, filename: str, headers: list[str], data: dict):
        self.filename = filename
        self.headers: list[str] = headers
        self.data: dict = data

    def _save_to(self, file):
        writer = csv.DictWriter(file, fieldnames=self.headers)

        writer.writeheader()

        for p in self.data:
            p_star = {key: value for key, value in p.items() if key in self.headers}

            writer.writerow(p_star)

    def get_iostream(self):
        result = io.BytesIO()
        self._save_to(result)

        return result.getvalue()

    def save(self, filepath):
        with open(filepath, "w") as file:
            self._save_to(file)


class CsvCreator:
    def __init__(self, provider: BaseProvider):
        self.provider: BaseProvider = provider
        self.populated_with_defaults: bool = False
        # The line below causes unique to fail because a
        # different faker is created each time, despite being for the
        # same generator.  Use singleton instead?
        self.faker: Faker = Faker("en_GB", generator=provider.generator)

    def get(self, headers: list[str], data: dict, filename: Optional[str]=None):
        headers = list(headers)
        filename = filename or self.faker.file_name(extension='csv')

        result = FakeCsvFile(filename, headers, data)

        return result


class FakeXlsFile():
    def __init__(self, filename: str, headers: list[str], data: dict):
        self.filename = filename
        self.headers: list[str] = headers
        self.data: dict = data

    def get_iostream(self):
        result = io.BytesIO()
        self.save(result)

        return result.getvalue()

    def save(self, filepath):
        wb = xlwt.Workbook()
        ws = wb.add_sheet('Test Sheet')
        style = xlwt.XFStyle()
        style.num_format_str = 'D-MMM-YY' # Other options: D-MMM-YY, D-MMM, MMM-YY, h:mm, h:mm:ss, h:mm, h:mm:ss, M/D/YY h:mm, mm:ss, [h]:mm:ss, mm:ss.0

        row_index = 0
        for col_index, h in enumerate(self.headers):
            ws.write(row_index, col_index, h)

        for row_index, p in enumerate(self.data, 1):
            for col_index, h in enumerate(self.headers):
                if h in p.keys():
                    ws.write(row_index, col_index, p[h], style)

        wb.save(filepath)


class XlsCreator:
    def __init__(self, provider: BaseProvider):
        self.provider: BaseProvider = provider
        self.populated_with_defaults: bool = False
        # The line below causes unique to fail because a
        # different faker is created each time, despite being for the
        # same generator.  Use singleton instead?
        self.faker: Faker = Faker("en_GB", generator=provider.generator)

    def get(self, headers: list[str], data: dict, filename: Optional[str]=None):
        headers = list(headers)
        filename = filename or self.faker.file_name(extension='csv')

        result = FakeXlsFile(filename, headers, data)

        return result


class LbrcFileProvider(BaseProvider):
    @cache
    def xlsx_file(self):
        return XlsxCreator(self)

    @cache
    def xls_file(self):
        return XlsCreator(self)

    @cache
    def csv_file(self):
        return CsvCreator(self)
